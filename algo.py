import string

from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np

inf = 999999


def floyd(file):
    result = comp_floyd(get_matrix(file))
    return result


def comp_floyd(incidence_matrix):
    x, y = incidence_matrix.shape
    dist = np.asarray(incidence_matrix)
    dist = np.where(dist == "M", inf, dist).astype(np.int)
    k_mat = np.zeros((x, y)).astype(np.int)
    result = dict()
    # Replace inf value 999999 with M for pretty print
    result["D 0"] = np.where(dist == 999999, "M", dist)
    for k in range(x):
        res_mat = np.copy(dist).astype(np.str)
        for i in range(x):
            for j in range(x):
                if dist[i][j] > dist[i][k] + dist[k][j]:
                    k_mat[i][j] = k + 1
                    res_mat[i][j] = np.str(dist[i][k] + dist[k][j]) + "*"
                dist[i][j] = min(dist[i][j], dist[i][k] + dist[k][j])

        result["D {}".format(k + 1)] = np.where(res_mat == "999999", "M", res_mat)
        # print(res_mat)

    result["K"] = k_mat
    return result


def get_matrix(file):
    if not isinstance(file, str):
        file = '/'.join(map(str, file))
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    distance_matrix = np.array([[value for value in row] for row in ws.values])
    return distance_matrix


def prim(file):
    dist_matrix = get_matrix(file)
    dist_matrix = np.where(dist_matrix == "M", inf, dist_matrix).astype(np.int)
    orig_matrix = np.copy(dist_matrix)

    if dist_matrix.shape[0] != dist_matrix.shape[1]:
        raise ValueError("Matica incidencii musi byt stvorcova")
    n_vertices = dist_matrix.shape[0]
    spanning_edges = []

    # initialize with node 0:
    visited_vertices = [0]
    num_visited = 1
    # exclude self connections:
    diag_indices = np.arange(n_vertices)
    dist_matrix[diag_indices, diag_indices] = inf

    while num_visited != n_vertices:
        new_edge = np.argmin(dist_matrix[visited_vertices], axis=None)
        # 2d encoding of new_edge from flat, get correct indices
        new_edge = divmod(new_edge, n_vertices)
        new_edge = [visited_vertices[new_edge[0]], new_edge[1]]
        # add edge to tree
        spanning_edges.append(new_edge)
        visited_vertices.append(new_edge[1])
        # remove all edges inside current tree
        dist_matrix[visited_vertices, new_edge[1]] = inf
        dist_matrix[new_edge[1], visited_vertices] = inf
        num_visited += 1
    mst = np.vstack(spanning_edges)
    result = [[]]
    weight = 0

    r_arr = [(edge, orig_matrix[edge[0]][edge[1]])for edge in mst]
    # for edge in mst:
    #     weight += orig_matrix[edge[0]][edge[1]]
    #     v1 = string.ascii_uppercase[edge[0]]
    #     v2 = string.ascii_uppercase[edge[1]]
    #     edge = v1 + "->" + v2
    #     result[0].append(edge)
    # result.append(["Vaha", weight])
    return r_arr


def minimal_addition(file):
    dist_matrix = get_matrix(file)
    dist_matrix = np.where(dist_matrix == "M", inf, dist_matrix).astype(int)

    size = dist_matrix.shape[0]
    change_matrix = np.zeros((size, size), dtype=int)

    # max number of vertices is 2^20
    s = 0
    for s in range(1, 20):
        if (size - 1) <= pow(2, s):
            break

    result = []
    mat_next = dist_matrix
    for k in range(s):
        mat = mat_next
        mat_next = np.copy(mat)
        for x in range(size):
            for y in range(size):
                dx = [mat[x][l] for l in range(size)]
                dy = [mat[l][y] for l in range(size)]
                sum_mat = np.add(dx, dy)
                min_value = np.amin(sum_mat)

                # get all indexes of min value as array and start from 1
                min_indexes = np.where(sum_mat == min_value)[0]
                min_indexes += 1
                if min_value < mat[x][y]:
                    mat_next[x][y] = min_value
                    change_matrix[x][y] = min_indexes[-1]

        def format_changes(value):
            if value == 0:
                return ""
            else:
                return "({})".format(value)

        res_mat = np.where(mat_next == inf, "M", mat_next).astype(str)
        fchange_mat = np.array([[format_changes(change_matrix[x][y]) for x in range(size)] for y in range(size)])

        result.append(np.char.add(res_mat, fchange_mat))
    return result


def minimal_addition_minimum_flow(file):
    dist_matrix = get_matrix(file)
    dist_matrix = np.where(dist_matrix == "M", 0, dist_matrix).astype(int)

    np.fill_diagonal(dist_matrix, 999999)

    size = dist_matrix.shape[0]
    change_matrix = np.zeros((size, size), dtype=int)

    # max number of vertices is 2^20
    s = 0
    for s in range(1, 20):
        if (size - 1) <= pow(2, s):
            break

    result = []
    mat_next = dist_matrix
    for k in range(s):
        mat = mat_next
        mat_next = np.copy(mat)
        for x in range(size):
            for y in range(size):
                dx = [mat[x][l] for l in range(size)]
                dy = [mat[l][y] for l in range(size)]

                max_arr = np.maximum(dx, dy)
                max_arr = max_arr[max_arr != 0]
                min_value = np.min(max_arr)

                min_indexes = np.where(max_arr == min_value)[0]
                min_indexes += 1

                # get all indexes of min value as array and start from 1
                min_indexes = np.where(max_arr == min_value)[0]
                min_indexes += 1
                if min_value < mat_next[x][y]:
                    mat_next[x][y] = min_value
                    change_matrix[x][y] = min_indexes[-1]

        def format_changes(value):
            if value == 0:
                return ""
            else:
                return "({})".format(value)

        res_mat = np.where(mat_next == inf, "M", mat_next).astype(str)
        fchange_mat = np.array([[format_changes(change_matrix[x][y]) for x in range(size)] for y in range(size)])

        result.append(np.char.add(res_mat, fchange_mat))
    return result


def minimal_addition_maximum_flow(file):
    dist_matrix = get_matrix(file)
    dist_matrix = np.where(dist_matrix == "M", 0, dist_matrix).astype(int)

    np.fill_diagonal(dist_matrix, inf)
    size = dist_matrix.shape[0]
    change_matrix = np.zeros((size, size), dtype=int)

    # max number of vertices is 2^20
    s = 0
    for s in range(1, 20):
        if (size - 1) <= pow(2, s):
            break

    result = []
    mat_next = dist_matrix
    for k in range(s):
        mat = mat_next
        mat_next = np.copy(mat)
        for x in range(size):
            for y in range(size):
                dx = [mat[x][l] for l in range(size)]
                dy = [mat[l][y] for l in range(size)]
                min_arr = np.minimum(dx, dy)
                max_value = np.max(min_arr)
                # get all indexes of min value as array and start from 1
                max_indexes = np.where(min_arr == max_value)[0]
                max_indexes += 1
                if max_value > mat[x][y]:
                    mat_next[x][y] = max_value
                    change_matrix[x][y] = max_indexes[0]

        def format_changes(value):
            if value == 0:
                return ""
            else:
                return "({})".format(value)

        res_mat = np.where(mat_next == inf, "M", mat_next).astype(str)
        fchange_mat = np.array([[format_changes(change_matrix[x][y]) for x in range(size)] for y in range(size)])

        result.append(np.char.add(res_mat, fchange_mat))
    return result

# return only matrix with shortest paths
def minimal_addition_array(file):
    dist_matrix = get_matrix(file)
    dist_matrix = np.where(dist_matrix == "M", inf, dist_matrix).astype(int)

    size = dist_matrix.shape[0]
    change_matrix = np.zeros((size, size), dtype=int)

    # max number of vertices is 2^20
    s = 0
    for s in range(1, 20):
        if (size - 1) < pow(2, s):
            break

    mat_next = dist_matrix
    for k in range(s):
        mat = mat_next
        mat_next = np.copy(mat)
        for x in range(size):
            for y in range(size):
                dx = [mat[x][l] for l in range(size)]
                dy = [mat[l][y] for l in range(size)]
                sum_mat = np.add(dx, dy)
                min_value = np.amin(sum_mat)

                if min_value < mat[x][y]:
                    mat_next[x][y] = min_value

    return mat_next


def shortest_paths(file):
    return


def best_neighbour_matrix(file):
    dist_matrix = minimal_addition_array(file)
    n = dist_matrix.shape[0]
    e_mat = np.zeros(dist_matrix.shape)
    for i in range(dist_matrix.shape[0]):
        for j in range(dist_matrix.shape[0]):
            dk = [dist_matrix[i][k] for k in range(dist_matrix.shape[0]) if k!=j]
            dq = [dist_matrix[q][j] for q in range(dist_matrix.shape[0]) if q!=i]
            e_mat[i][j] = (n - 2) * dist_matrix[i][j] - np.sum(dk) - np.sum(dq)

    np.fill_diagonal(e_mat, 0)
    return e_mat


def best_neighbour(file, initial):
    dist_matrix = minimal_addition_array(file)
    e_mat = best_neighbour_matrix(file)
    mark_matrix = np.copy(e_mat).astype(str)
    n = e_mat.shape[0]
    k = 1
    i = initial
    i_arr = []
    j = [x for x in range(n) if x != i]
    while k < n:
        row = [(e_mat[i][ji], ji) for ji in j]
        min_index = min(row, key=lambda t: t[0])[1]
        i_arr.append((i, dist_matrix[i][min_index]))
        mark_matrix[i][min_index] += "*"
        i = min_index
        j.remove(min_index)
        k += 1
    i_arr.append((i, dist_matrix[i][initial]))
    return [i_arr, mark_matrix]


# returns list with tuple and array [(Verticle, Weight), [Visualization Array]]
def nearest_neighbour(file, initial):
    dist_matrix = minimal_addition_array(file)
    mark_matrix = np.copy(dist_matrix).astype(str)
    n = dist_matrix.shape[0]
    k = 1
    i = initial
    i_arr = []
    j = [x for x in range(n) if x != i]
    while k < n:
        row = [(dist_matrix[i][ji], ji) for ji in j]
        min_index = min(row, key=lambda t: t[0])[1]
        i_arr.append((i, dist_matrix[i][min_index]))
        mark_matrix[i][min_index] += "*"
        i = min_index
        j.remove(min_index)
        k += 1
    i_arr.append((i, dist_matrix[i][initial]))
    return [i_arr, mark_matrix]


def get_tasks(file):
    if not isinstance(file, str):
        file = '/'.join(map(str, file))
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    # as object to allow addition of lists to array element
    tasks_matrix = np.array([[value for value in row] for row in ws.values]).astype(object)

    # split all second columns (next tasks) to list
    tasks_matrix[:, 1] = [tasks.split(",") for tasks in tasks_matrix[:, 1]]

    return tasks_matrix


def get_tasks_graph(tasks_matrix):
    nodes = tasks_matrix[:, 0]
    edges = []
    #print(tasks_matrix[:, 0:2])
    for row in tasks_matrix[:, 0:2]:
        for task in row[1]:
            if task is not "-":
                #edges.append([row[0], task])
                edges.append([task, row[0]])

    return nodes, edges


def cpm(file):
    import arrowGraphGenerator as agg
    tasks = get_tasks(file)
    nodes, edges = agg.generate_arrow_graph(tasks)
    #tasks = tasks.tolist()
    tasks = list(tasks)


    dummy_edges = [edge for edge in edges if edge[2]['name'] == '']
    full_edges = [edge for edge in edges if edge[2]['name'] != '']
    # print(full_edges)

    # for edge in full_edges:
    #     edge[2]['name'] = string.ascii_uppercase[int(edge[2]['name'])]

    # Collapse end nodes if possible
    end_nodes = [edge[1] for edge in edges]
    start_nodes = [edge[0] for edge in edges]

    collapse_nodes = [node for node in end_nodes if node not in start_nodes]
    end_nodes = [node for node in end_nodes if node in start_nodes]

    nodes = [node for node in nodes if node not in collapse_nodes[1:]]

    for edge in full_edges:
        if edge[1] in collapse_nodes:
            edge[1] = collapse_nodes[0]

    # Collapse start nodes if possible
    # all end nodes
    end_nodes = [edge[1] for edge in edges]
    # all start nodes
    start_nodes = [edge[0] for edge in edges]

    collapse_nodes = [node for node in start_nodes if node not in end_nodes]
    start_nodes = [node for node in start_nodes if node in end_nodes]

    # remove nodes that can be collapsed but leave one
    nodes = [node for node in nodes if node not in collapse_nodes[1:]]

    for edge in full_edges:
        if edge[0] in collapse_nodes:
            edge[0] = collapse_nodes[0]

    dummy_edges = [edge for edge in edges if edge[2]['name'] is '']

    letters = string.ascii_uppercase
    # for index, edge in enumerate(dummy_edges):
        # tasks.append(["FC{}".format(index+1), [edge[0], edge[1]], '0'])

    for task in tasks:
        task[1] = [[edge[0], edge[1]] for edge in edges if edge[2]['name'] == task[0]]
    for task in tasks:
        print(task)
    #print(tasks)


cpm("input/cpm2.xlsx")